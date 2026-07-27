from openpyxl import Workbook,load_workbook

# Create Workbook
wb = Workbook()

# Active Sheet
sheet = wb.active
sheet.title = "Selling Report" 

# Header
sheet.append(["Product", "Number", "Price", "Total"])

# Data
sell = [
    ["Shirt", 3, 2000],
    ["Pant", 4, 3000],
    ["T-shirts", 10, 3000],
]

total_sell = 0

#Calculate Total

for product, number, price in sell:
    total = number * price
    sheet.append([product, number, price, total])
    total_sell += total

# Grand Total
sheet.append(["", "", "Total Sell:", total_sell])

# Save Excel File & see how many rows

wb.save("sell_report.xlsx")

print(f"Report Total Sell: {total_sell} Tk")

wb = load_workbook("sell_report.xlsx")
sheet = wb.active

for row in sheet.iter_rows(values_only=True):
    print(row)

import os 
print(os.getcwd())
print(os.listdir())

from pathlib import Path

path = Path.cwd()
print(path)

file = Path("sell_report.xlsx") 
print(file.exists())



