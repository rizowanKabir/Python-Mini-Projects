from pathlib import Path 
from openpyxl import load_workbook
from PyPDF2 import PdfMerger
from PIL import Image
import schedule
import time
from datetime import datetime

# Create Required Folders

BASE_DIR = Path.cwd()

reports = BASE_DIR / "reports"
invoices = BASE_DIR / "invoices"
images = BASE_DIR / "images"
backup = BASE_DIR / "backup"

for folder in [reports, invoices, images, backup]:
    folder.mkdir(exist_ok=True)


def automation():

    print("Automation Started...\n")

    today = datetime.today().strftime("%Y-%m-%d")

    backup_folder = backup / today
    backup_folder.mkdir(exist_ok=True)

    # Excel Automation

    excel_file = reports / "sales.xlsx"

    if excel_file.exists():

        wb = load_workbook(excel_file)
        sheet = wb.active

        grand_total = 0

        if sheet.cell(row=1, column=4).value != "Total":
            sheet.cell(row=1, column=4).value = "Total"

        for row in range(2, sheet.max_row + 1):

            qty = sheet.cell(row=row, column=2).value
            price = sheet.cell(row=row, column=3).value

            total = qty * price

            sheet.cell(row=row, column=4).value = total

            grand_total += total

        sheet.cell(row=sheet.max_row + 1, column=3).value = "Grand Total"
        sheet.cell(row=sheet.max_row, column=4).value = grand_total

        output_excel = backup_folder / "sales_report.xlsx"

        wb.save(output_excel)

        print("Excel Done")

    else:
        print("sales.xlsx not found!")

    # Merge PDF

    merger = PdfMerger()

    pdf_files = invoices.glob("*.pdf")

    count = 0

    for pdf in pdf_files:
        merger.append(str(pdf))
        count += 1

    if count > 0:

        merged_pdf = backup_folder / "all_invoice.pdf"

        merger.write(str(merged_pdf))
        merger.close()

        print("PDF Merge Done")

    else:
        print("No PDF Found")

    # Resize Images

    image_backup = backup_folder / "images"
    image_backup.mkdir(exist_ok=True)

    extensions = ["*.jpg", "*.jpeg", "*.png"]

    image_count = 0

    for ext in extensions:

        for image_path in images.glob(ext):

            img = Image.open(image_path)

            img = img.resize((800, 800))

            img.save(image_backup / image_path.name)

            image_count += 1

    print(f"{image_count} Images Resized")

    print("\nAutomation Finished")

# Scheduler

schedule.every().day.at("09:00").do(automation)

print("Scheduler Running...")

while True:

    schedule.run_pending()

    time.sleep(1)
